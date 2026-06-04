# -*- coding: utf-8 -*-
import openpyxl, json, os, re, unicodedata

SRC = r'c:/Users/TRIA 2026/Downloads/Loovi - Ressarcimento/2025 - Loovi Ressarcimento - ORGANIZADO.xlsx'
OUT = r'c:/Users/TRIA 2026/Downloads/Loovi - Ressarcimento/loovi-web/data'
os.makedirs(os.path.join(OUT, 'sheets'), exist_ok=True)

# index in source row -> (short key, truncate length)
COLS = {
    0:  ('arquivo', 200),
    1:  ('pessoa', 80),
    6:  ('placa', 12),
    8:  ('veiculo', 80),
    9:  ('titular', 40),
    15: ('data_fato', 30),
    17: ('dia_semana', 20),
    20: ('num_bo', 40),
    24: ('fin_tipo', 30),
    25: ('coberto', 12),
    27: ('cnh_status', 20),
    31: ('cnh_pontos', 8),
    47: ('resumo', 600),
    49: ('conclusao', 400),
    51: ('classif', 30),
    52: ('monta', 16),
    53: ('analista', 40),
    54: ('inicio', 30),
    55: ('fim', 30),
}

def trunc(v, n):
    s = '' if v is None else str(v).strip()
    if s.lower() in ('none', 'nan', 'null'):
        s = ''
    return s if len(s) <= n else s[:n].rstrip() + '…'

def slugify(name):
    s = unicodedata.normalize('NFKD', name).encode('ascii', 'ignore').decode('ascii')
    s = s.lower()
    s = re.sub(r'[^a-z0-9]+', '-', s).strip('-')
    return s

def parse_title(title):
    m = re.match(r'^(.*?)\s*-\s*(20\d{2})$', title)
    if m:
        return m.group(1).strip(), m.group(2)
    return title.strip(), ''

wb = openpyxl.load_workbook(SRC, read_only=True)

# ---- per-sheet data + index ----
index = []
for ws in wb.worksheets:
    if ws.title == 'RESUMO':
        continue
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        continue
    body = [r for r in rows[1:] if any(c is not None and str(c).strip() != '' for c in r)]
    if not body:
        # still register empty sheets that are not pure noise
        pass
    analyst, year = parse_title(ws.title)
    slug = slugify(ws.title)
    data = []
    for r in body:
        obj = {}
        for i, (key, n) in COLS.items():
            val = trunc(r[i], n) if i < len(r) else ''
            if val:
                obj[key] = val
        data.append(obj)
    with open(os.path.join(OUT, 'sheets', slug + '.json'), 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, separators=(',', ':'))
    index.append({
        'slug': slug,
        'sheet': ws.title,
        'analyst': analyst,
        'year': year,
        'count': len(data),
    })

index.sort(key=lambda x: (x['analyst'], x['year']))
with open(os.path.join(OUT, 'index.json'), 'w', encoding='utf-8') as f:
    json.dump(index, f, ensure_ascii=False, indent=0)

# ---- summary (from RESUMO sheet) ----
ws = wb['RESUMO']
rows = list(ws.iter_rows(values_only=True))
analysts = []
for r in rows[3:]:  # data starts row 4 (0-indexed 3)
    name = r[0]
    if not name:
        continue
    name = str(name).strip()
    if name.upper() in ('TOTAL',):
        continue
    if name.upper().startswith('REVISAR') or name.upper().startswith('TOTAL GERAL'):
        continue
    try:
        y24 = int(r[1] or 0); y25 = int(r[2] or 0); tot = int(r[3] or 0)
    except (ValueError, TypeError):
        continue
    analysts.append({'analyst': name, 'y2024': y24, 'y2025': y25, 'total': tot})

total_docs = sum(a['total'] for a in analysts)
total_24 = sum(a['y2024'] for a in analysts)
total_25 = sum(a['y2025'] for a in analysts)
analysts.sort(key=lambda x: -x['total'])

summary = {
    'total_docs': total_docs,
    'total_2024': total_24,
    'total_2025': total_25,
    'n_analysts': len([a for a in analysts if a['total'] > 0]),
    'n_sheets': len(index),
    'analysts': analysts,
}
with open(os.path.join(OUT, 'summary.json'), 'w', encoding='utf-8') as f:
    json.dump(summary, f, ensure_ascii=False, indent=0)

print('OK')
print('sheets:', len(index), '| total docs:', total_docs, '| 2024:', total_24, '| 2025:', total_25)
tot_size = sum(os.path.getsize(os.path.join(OUT, 'sheets', f)) for f in os.listdir(os.path.join(OUT, 'sheets')))
print('data/sheets total size:', round(tot_size/1024/1024, 1), 'MB')
