# -*- coding: utf-8 -*-
import openpyxl, json, re, os, time, urllib.request, urllib.error

SRC = r'c:/Users/TRIA 2026/Downloads/Loovi - Ressarcimento/2025 - Loovi Ressarcimento - ORGANIZADO.xlsx'
SB_URL = os.environ['SB_URL']
SB_SECRET = os.environ['SB_SECRET']
TABLE = 'loovi_ressarcimento'
BATCH = 500

def parse_title(t):
    m = re.match(r'^(.*?)\s*-\s*(20\d{2})$', t)
    if m:
        return m.group(1).strip(), int(m.group(2))
    return t.strip(), None

def clean(v):
    if v is None:
        return None
    s = str(v).strip()
    if s == '' or s.lower() in ('none', 'nan', 'null'):
        return None
    return s

def post(rows):
    body = json.dumps(rows, ensure_ascii=False).encode('utf-8')
    req = urllib.request.Request(
        f'{SB_URL}/rest/v1/{TABLE}',
        data=body, method='POST',
        headers={
            'apikey': SB_SECRET,
            'Authorization': f'Bearer {SB_SECRET}',
            'Content-Type': 'application/json',
            'Prefer': 'return=minimal',
        })
    for attempt in range(4):
        try:
            with urllib.request.urlopen(req, timeout=120) as r:
                return r.status
        except urllib.error.HTTPError as e:
            msg = e.read().decode('utf-8', 'replace')[:300]
            if attempt == 3:
                raise SystemExit(f'HTTP {e.code} on batch: {msg}')
            time.sleep(2 * (attempt + 1))
        except Exception as e:
            if attempt == 3:
                raise SystemExit(f'ERR: {e}')
            time.sleep(2 * (attempt + 1))

wb = openpyxl.load_workbook(SRC, read_only=True)
# fixed header/key set (same schema across all sheets) -> uniform keys per batch
hdr = [c.value for c in next(wb['ALISSON COSTA - 2025'].iter_rows(min_row=1, max_row=1))]
ALLKEYS = ['sheet', 'analista_tab', 'ano'] + list(hdr)
buf = []
sent = 0
for ws in wb.worksheets:
    if ws.title == 'RESUMO':
        continue
    analista_tab, ano = parse_title(ws.title)
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if not any(c is not None and str(c).strip() != '' for c in r):
            continue
        obj = {k: None for k in ALLKEYS}
        obj['sheet'] = ws.title
        obj['analista_tab'] = analista_tab
        obj['ano'] = ano
        for h, v in zip(hdr, r):
            cv = clean(v)
            if cv is not None:
                obj[h] = cv
        buf.append(obj)
        if len(buf) >= BATCH:
            post(buf)
            sent += len(buf)
            print(f'  sent {sent}', flush=True)
            buf = []
if buf:
    post(buf)
    sent += len(buf)
print('DONE. total rows sent:', sent)
